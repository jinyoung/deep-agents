"""
Deep Agents + Anthropic xlsx Skill + Docker Sandbox 예제

Docker 컨테이너 안에서 에이전트가 직접 Python 코드를 실행하여
엑셀 파일을 생성합니다. xlsx 스킬의 recalc.py도 자동 호출됩니다.

사전 준비:
  docker build -t deepagents-xlsx .
  docker run -d --name deepagents-sandbox deepagents-xlsx
"""

import os
import subprocess

from dotenv import load_dotenv

from deepagents import create_deep_agent
from langchain_core.messages import HumanMessage

from docker_sandbox import DockerSandboxBackend

load_dotenv()

CONTAINER_NAME = os.environ.get("CONTAINER_NAME", "deepagents-sandbox")

# Docker 샌드박스 백엔드 생성
backend = DockerSandboxBackend(
    container_name=CONTAINER_NAME,
    workdir="/workspace",
)

# xlsx 스킬을 포함한 Deep Agent 생성
agent = create_deep_agent(
    model="openai:gpt-4o",
    skills=["skills/xlsx/"],
    backend=backend,
    system_prompt=(
        "당신은 엑셀 파일 생성 전문가입니다. "
        "사용자의 요청에 따라 execute 도구로 Python 코드를 직접 실행하여 엑셀 파일을 만드세요. "
        "반드시 Excel 수식을 사용하고, Python에서 값을 계산하여 하드코딩하지 마세요. "
        "openpyxl 라이브러리가 이미 설치되어 있습니다. "
        "수식이 포함된 엑셀 파일을 만든 후에는 반드시 "
        "python /workspace/skills/xlsx/scripts/recalc.py <파일경로> 를 실행하여 "
        "수식 값을 재계산하고 오류를 검증하세요. "
        "파일은 /workspace/ 디렉토리에 저장하세요."
    ),
)

# 에이전트 실행
print(">>> Deep Agent에게 엑셀 생성 요청 중 (Docker Sandbox)...")
result = agent.invoke(
    {
        "messages": [
            HumanMessage(
                content=(
                    "2024년 분기별 매출 보고서 엑셀 파일을 만들어주세요.\n\n"
                    "요구사항:\n"
                    "- 제품: 스마트폰(150,200,180,220), 노트북(300,280,320,350), 태블릿(80,90,85,100) (단위: 억원)\n"
                    "- 헤더: 제품명, Q1, Q2, Q3, Q4, 연간합계\n"
                    "- 각 제품의 연간합계는 =SUM() 수식 사용\n"
                    "- 하단 분기별/연간 전체합계도 =SUM() 수식 사용\n"
                    "- 헤더는 볼드체 + 회색 배경 + 중앙정렬\n"
                    "- 숫자 셀은 천 단위 쉼표 포맷 적용\n"
                    "- 열 너비 적절히 조정\n"
                    "- 파일명: /workspace/sales_report_2024.xlsx\n\n"
                    "execute 도구로 Python 코드를 직접 실행하여 파일을 생성하고, "
                    "생성 후 python /workspace/skills/xlsx/scripts/recalc.py /workspace/sales_report_2024.xlsx 를 "
                    "실행하여 수식을 재계산하고 오류를 검증해주세요."
                )
            )
        ]
    }
)

# 에이전트 응답 출력
final_message = result["messages"][-1]
print("\n=== Agent 응답 ===")
content = final_message.content if hasattr(final_message, "content") else str(final_message)
if isinstance(content, list):
    for item in content:
        if isinstance(item, dict) and item.get("type") == "text":
            print(item["text"])
else:
    print(content)

# 컨테이너에서 파일을 로컬로 복사
print("\n>>> 컨테이너에서 엑셀 파일 복사 중...")
os.makedirs("output", exist_ok=True)
cp_result = subprocess.run(
    ["docker", "cp", f"{CONTAINER_NAME}:/workspace/sales_report_2024.xlsx", "output/"],
    capture_output=True,
    text=True,
)

if cp_result.returncode == 0:
    print(">>> 파일 복사 완료: output/sales_report_2024.xlsx")

    # 엑셀 파일 내용 검증
    from openpyxl import load_workbook

    wb = load_workbook("output/sales_report_2024.xlsx")
    ws = wb.active
    print(f"    시트명: {ws.title}")
    print(f"    범위: {ws.dimensions}\n")
    print("    --- 셀 내용 ---")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        line = []
        for cell in row:
            if cell.value is not None:
                line.append(f"{cell.coordinate}={cell.value}")
        if line:
            print(f"    {', '.join(line)}")
else:
    print(f">>> 파일 복사 실패: {cp_result.stderr}")
